import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

from package_for_automation.functions import copy_sheet
from package_for_automation.functions import set_border_for_inner_cells
from package_for_automation.functions import set_border_for_last_row
from package_for_automation.functions import set_border_for_last_column
from package_for_automation.functions import cell_alignment

def excel_file_automate(
        original, template, saveTo, output_quantity, file_name
):

    # Record the start time
    start_time = time.time()

    print("In progress..")

    counter_for_testing = 0

    source_filename = original
    source_workbook = openpyxl.load_workbook(source_filename)
    source_sheet = source_workbook.active

    name_rows = {}

    for row in source_sheet.iter_rows(
        min_row=2, max_row=source_sheet.max_row, values_only=True
    ):
        name = row[2]
        if name not in name_rows:
            name_rows[name] = []
        name_rows[name].append(row)

    for name, rows in name_rows.items():
        target_filename = f'{name}_{file_name}.xlsx'
        target_filename_without_backslash = target_filename.replace("/", "-")
        target_workbook = openpyxl.Workbook()
        target_sheet = target_workbook.active

        # Copy rows
        start_row = 8
        for i, row in enumerate(rows, start=start_row):
            for j, value in enumerate(row, start=1):
                target_sheet.cell(row=i, column=j, value=value)

        target_workbook.save(f'{saveTo}/{target_filename_without_backslash}')
        target_workbook.close

        # Move rows in correct place
        reopen_filename = f'{saveTo}/{target_filename_without_backslash}'
        rearrange_cols_workbook = load_workbook(reopen_filename)
        work_sheet = rearrange_cols_workbook['Sheet']
        work_sheet.move_range(
            cell_range=f'D8:O{work_sheet.max_row}', rows=0, cols=-2
        )

        # Copy template into file
        template_source_wb_for_copy = load_workbook(template, data_only=True)
        source_sheet = template_source_wb_for_copy.active
        copy_sheet(source_sheet, work_sheet)

        # Format cells for template, freeze panes and make "SUM"
        work_sheet['B3'] = f'{name}'
        cell_formatting = work_sheet['B3']
        cell_formatting.font = Font(bold=True, name='Arial', size=16)
        cell_formatting.alignment = Alignment(
            horizontal='left', vertical='center'
        )
        work_sheet.freeze_panes = 'A8'
        work_sheet[f'J{work_sheet.max_row + 1}'] = (
            f'=SUM(J8:J{work_sheet.max_row})'
        )
        
        # Currency format set
        cell_count = 8
        while cell_count < work_sheet.max_row + 1:
            currency_cell = work_sheet[f'J{cell_count}']
            currency_cell.number_format = '#,##0 Ft'
            cell_count += 1

        # Set font and wrap text for all the cells
        cell_alignment(work_sheet, f'A8:M{work_sheet.max_row}')

        # Set font of the 'dates'
        for row in work_sheet[f'A8:A{work_sheet.max_row - 1}']:
            for cell in row:
                cell.font = Font(name='Arial', size=9, bold=True)

        # Set alignement for HUF row
        for row in work_sheet[f'J8:J{work_sheet.max_row}']:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='right', vertical='center'
                )

        # Write "osszesen"
        work_sheet[f'B{work_sheet.max_row}'] = "ÖSSZESEN"
        
        # Last row format (bold)
        for row in work_sheet[f'A{work_sheet.max_row}:M{work_sheet.max_row}']:
            for cell in row:
                cell.font = Font(bold=True, name='Arial')

        # Sat border (keret)
        set_border_for_inner_cells(work_sheet, f'A8:L{work_sheet.max_row}')
        set_border_for_last_row(
            work_sheet, f'A{work_sheet.max_row}:L{work_sheet.max_row}'
        )
        set_border_for_last_column(
            work_sheet, f'M8:M{work_sheet.max_row - 1}'
        )

        cell_border_last_cell = work_sheet.cell(
            row=work_sheet.max_row, column=13
        )
        cell_border_last_cell.border = Border(
            top=Side(border_style='medium',color='FF000000'),
            bottom=Side(border_style='medium',color='FF000000'), 
            right=Side(border_style='medium',color='FF000000')
        )
        
        # Set printing
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(
            work_sheet, 
            paper_size = 1, 
            orientation='landscape'
        )
        work_sheet.sheet_properties.pageSetUpPr.fitToPage = True 
        work_sheet.page_setup.fitToHeight = False

        # Set row diemensions
        for row in range(8, work_sheet.max_row + 1): 
            # For some reason max_row does not work properly,
            # so I had to do "+1"
            work_sheet.row_dimensions[row].height = 21.75

        rearrange_cols_workbook.active.sheet_view.zoomScale = 73
        rearrange_cols_workbook.save(
            f'{saveTo}/{target_filename_without_backslash}'
        )
        
        # For testing:
        counter_for_testing += 1
        if counter_for_testing == output_quantity:
            break

    source_workbook.close()

    # Record the end time
    end_time = time.time()

    print("Finished!")

    # Calculate and print the elapsed time
    elapsed_time = end_time - start_time
    elapsed_time_in_minutes = elapsed_time / 60
    print(f"Elapsed time: {elapsed_time} seconds")